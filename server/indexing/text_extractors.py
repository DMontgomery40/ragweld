from __future__ import annotations

import csv
from pathlib import Path


def extract_text_for_path(
    path: Path,
    *,
    parquet_max_rows: int = 5000,
    parquet_max_chars: int = 2_000_000,
    parquet_max_cell_chars: int = 20_000,
    parquet_text_columns_only: bool = True,
    parquet_include_column_names: bool = True,
) -> str | None:
    """Return extracted text for a file, or None if unsupported/unreadable.

    This is intentionally best-effort and dependency-light:
    - Text formats are read as UTF-8 (errors ignored)
    - PDF extraction uses pypdf if installed
    - XLSX extraction uses openpyxl if installed
    - Parquet extraction uses pyarrow if installed (bounded by config)
    """
    ext = path.suffix.lower()
    if ext in {".txt", ".md", ".rst", ".json", ".yaml", ".yml", ".toml", ".sql", ".py", ".js", ".jsx", ".ts", ".tsx"}:
        return _read_text(path)
    if ext in {".csv", ".tsv"}:
        return _read_delimited(path, delimiter="," if ext == ".csv" else "\t")
    if ext == ".pdf":
        return _read_pdf(path)
    if ext == ".xlsx":
        return _read_xlsx(path)
    if ext == ".parquet":
        return _read_parquet(
            path,
            max_rows=int(parquet_max_rows),
            max_chars=int(parquet_max_chars),
            max_cell_chars=int(parquet_max_cell_chars),
            text_columns_only=bool(parquet_text_columns_only),
            include_column_names=bool(parquet_include_column_names),
        )
    return None


def _read_text(path: Path) -> str | None:
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return None


def _read_delimited(path: Path, *, delimiter: str) -> str | None:
    try:
        raw = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return None

    # Normalize into a simple “table-ish” textual representation.
    out_lines: list[str] = []
    try:
        reader = csv.reader(raw.splitlines(), delimiter=delimiter)
        for row in reader:
            if not row:
                continue
            out_lines.append("\t".join(str(c).strip() for c in row))
    except Exception:
        return raw
    return "\n".join(out_lines)


def _read_pdf(path: Path) -> str | None:
    try:
        from pypdf import PdfReader
    except Exception:
        return None

    try:
        reader = PdfReader(str(path))
    except Exception:
        return None

    parts: list[str] = []
    for i, page in enumerate(getattr(reader, "pages", []) or []):
        try:
            txt = page.extract_text() or ""
        except Exception:
            txt = ""
        if not txt.strip():
            continue
        parts.append(f"\n\n--- page {i + 1} ---\n\n{txt.strip()}\n")
    joined = "\n".join(parts).strip()
    return joined or ""


def _read_xlsx(path: Path) -> str | None:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:
        return None

    out_lines: list[str] = []
    try:
        for ws in wb.worksheets:
            title = str(getattr(ws, "title", "") or "").strip() or "Sheet"
            out_lines.append(f"\n\n--- sheet {title} ---\n")
            try:
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    cells = [("" if c is None else str(c)).strip() for c in row]
                    if not any(cells):
                        continue
                    out_lines.append("\t".join(cells))
            except Exception:
                continue
    finally:
        try:
            wb.close()
        except Exception:
            pass

    joined = "\n".join(out_lines).strip()
    return joined or ""


def _read_parquet(
    path: Path,
    *,
    max_rows: int,
    max_chars: int,
    max_cell_chars: int,
    text_columns_only: bool,
    include_column_names: bool,
) -> str | None:
    """Read a Parquet file into a bounded text representation.

    This is designed for indexing: avoid loading huge Parquet files into memory.
    """
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except Exception:
        return None

    max_rows = max(1, int(max_rows))
    max_chars = max(1, int(max_chars))
    max_cell_chars = max(1, int(max_cell_chars))

    def _is_text_type(t: pa.DataType) -> bool:
        if pa.types.is_string(t) or pa.types.is_large_string(t):
            return True
        if pa.types.is_dictionary(t):
            try:
                return _is_text_type(t.value_type)
            except Exception:
                return False
        return False

    try:
        pf = pq.ParquetFile(str(path))
    except Exception:
        return None

    cols: list[str] | None = None
    if text_columns_only:
        try:
            schema = pf.schema_arrow
            text_cols = [f.name for f in schema if _is_text_type(f.type)]
            cols = text_cols or None
        except Exception:
            cols = None

    out_parts: list[str] = []
    total_rows = 0
    total_chars = 0

    try:
        for batch in pf.iter_batches(batch_size=1024, columns=cols):
            if total_rows >= max_rows or total_chars >= max_chars:
                break
            table = pa.Table.from_batches([batch])
            batch_rows = int(table.num_rows or 0)
            if batch_rows <= 0:
                continue

            names = list(table.column_names)
            arrays = [table.column(i).to_pylist() for i in range(table.num_columns)]
            for i in range(batch_rows):
                if total_rows >= max_rows or total_chars >= max_chars:
                    break
                row_parts: list[str] = []
                for col_name, col_vals in zip(names, arrays, strict=True):
                    try:
                        v = col_vals[i]
                    except Exception:
                        continue
                    if v is None:
                        continue
                    s = str(v).strip()
                    if not s:
                        continue
                    if len(s) > max_cell_chars:
                        s = s[:max_cell_chars] + "…"
                    row_parts.append(f"[{col_name}]\n{s}" if include_column_names else s)

                if row_parts:
                    chunk = f"\n\n--- row {total_rows} ---\n\n" + "\n\n".join(row_parts)
                    out_parts.append(chunk)
                    total_chars += len(chunk)
                total_rows += 1
    except Exception:
        return None

    joined = "\n".join(out_parts).strip()
    if total_rows >= max_rows and len(joined) < max_chars:
        joined = (joined + "\n\n… (truncated by parquet_extract_max_rows)\n").strip()
    elif total_chars >= max_chars:
        joined = (joined[:max_chars] + "\n\n… (truncated by parquet_extract_max_chars)\n").strip()
    return joined or ""
